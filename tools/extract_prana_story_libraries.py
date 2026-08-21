from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.replace("\r", " ").replace("\n", " ").strip()
        value = re.sub(r"\s+", " ", value)
        return value or None
    return value


def split_multi(value):
    text = clean(value)
    if not text:
        return []
    parts = re.split(r"\s*[•,]\s*|\s*;\s*", text)
    return [part.strip() for part in parts if part and part.strip() and part.strip() != "—"]


def split_arrows(value):
    text = clean(value)
    if not text:
        return []
    parts = re.split(r"\s*→\s*|\s*->\s*", text)
    return [part.strip() for part in parts if part and part.strip()]


def slugify(text):
    text = clean(text) or ""
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def load_table(ws, header_name):
    header_row = None
    headers = None
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cleaned = [clean(cell) for cell in row]
        if header_name in cleaned:
            header_row = idx
            headers = cleaned
            break
    if header_row is None:
        raise ValueError(f"Could not find header {header_name!r} in {ws.title}")

    data = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        cleaned = [clean(cell) for cell in row]
        if not any(cleaned):
            continue
        if cleaned[0] and isinstance(cleaned[0], str) and cleaned[0].startswith("Batch "):
            continue
        if cleaned[0] and isinstance(cleaned[0], str) and re.match(r"^[A-Z]\.\s", cleaned[0]):
            continue
        if cleaned == headers:
            continue
        entry = {}
        for key, value in zip(headers, cleaned):
            if key:
                entry[key] = value
        data.append(entry)
    return data


def parse_childhood_situations(ws):
    rows = load_table(ws, "#")
    items = []
    for row in rows:
        if not row.get("Situation"):
            continue
        item = {
            "id": int(row["#"]),
            "slug": slugify(row["Situation"]),
            "name": row["Situation"],
            "category": row.get("Category"),
            "coreNeed": row.get("Core Need"),
            "falseBelief": row.get("False Belief"),
            "trueBelief": row.get("True Belief"),
            "feeling": row.get("What the Child is Feeling"),
            "parentTeach": row.get("What the Parent Hopes to Teach"),
            "wisdomElement": row.get("Primary Symbol"),
            "secondaryWisdomElement": row.get("Secondary Symbol"),
            "lifeDomains": split_multi(row.get("Action Domains")),
            "priority": row.get("Production Priority"),
        }
        items.append(item)
    return items


def parse_named_rows(ws, first_header, key_field, field_map, split_fields=None):
    split_fields = split_fields or {}
    rows = load_table(ws, first_header)
    items = []
    for index, row in enumerate(rows, start=1):
        key_value = row.get(key_field)
        if not key_value or not isinstance(key_value, str):
            continue
        item = {"id": index, "slug": slugify(key_value), "name": key_value}
        has_payload = False
        for src, dest in field_map.items():
            value = row.get(src)
            if src in split_fields:
                parser = split_fields[src]
                value = parser(value)
            if value not in (None, [], ""):
                has_payload = True
            item[dest] = value
        if not has_payload:
            continue
        items.append(item)
    return items


def parse_adventure_triggers(ws):
    items = []
    current_archetype = None
    for row in ws.iter_rows(values_only=True):
        first = clean(row[0] if len(row) > 0 else None)
        second = clean(row[1] if len(row) > 1 else None)
        if not first and not second:
            continue
        if first and isinstance(first, str) and re.match(r"^[A-Z].*[A-Za-z]$", first) and not second:
            if first not in {"ADVENTURE TRIGGER LIBRARY (LOCKED — 55 × 8 = 440)", "RESCUE & PROTECTION"} and not re.match(r"^[A-Z]\.\s", first):
                current_archetype = first
            continue
        if isinstance(first, int) and second:
            items.append(
                {
                    "id": len(items) + 1,
                    "archetype": current_archetype,
                    "name": second,
                }
            )
    return items


def parse_scorecard(ws):
    categories = []
    current = None
    for row in ws.iter_rows(values_only=True):
        first = clean(row[0] if len(row) > 0 else None)
        if not first:
            continue
        match = re.match(r"^\d+\.\s+(.*)\s+\(/(\d+)\)$", first)
        if match:
            current = {"name": match.group(1), "maxScore": int(match.group(2)), "checks": []}
            categories.append(current)
            continue
        if current and first.startswith("☐"):
            current["checks"].append(first.replace("☐", "").strip())
    return categories


def build_bundle(workbook_path):
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)

    libraries = {
        "childhoodSituations": parse_childhood_situations(wb["Childhood Situation Library"]),
        "wisdomElements": parse_named_rows(
            wb["Wisdom Element Library"],
            "Wisdom Element",
            "Wisdom Element",
            {
                "Type": "type",
                "Core Meaning": "coreMeaning",
                "Best Core Needs": "bestCoreNeeds",
                "Best Childhood Challenges": "bestChallenges",
                "Teaching Method": "teachingMethod",
                "Story Expression": "storyExpression",
                "Symbol-inspired Insight": "insight",
                "Best Emotional Arc": "bestEmotionalArcs",
                "Best Story Worlds": "bestStoryWorlds",
            },
            {
                "Best Core Needs": split_multi,
                "Best Childhood Challenges": split_multi,
                "Best Emotional Arc": split_multi,
                "Best Story Worlds": split_multi,
            },
        ),
        "mainCharacters": parse_named_rows(
            wb["Main Character Library"],
            "Character",
            "Character",
            {
                "Core Personality": "corePersonality",
                "Can Also Be": "canAlsoBe",
                "Primary Story Role": "primaryStoryRole",
                "Best Core Needs": "bestCoreNeeds",
                "Best Childhood Challenges": "bestChallenges",
                "Best Ganesha Symbols": "bestWisdomElements",
            },
            {
                "Can Also Be": split_multi,
                "Best Core Needs": split_multi,
                "Best Childhood Challenges": split_multi,
                "Best Ganesha Symbols": split_multi,
            },
        ),
        "lifeDomains": parse_named_rows(
            wb["Life Domain Library"],
            "Life Domain",
            "Life Domain",
            {
                "Simple Description": "description",
                "Typical Story Focus": "storyFocus",
                "Example Missions": "exampleMissions",
            },
            {"Example Missions": split_multi},
        ),
        "storyActions": parse_named_rows(
            wb["Story Action Library"],
            "Story Action",
            "Story Action",
            {
                "Simple Description": "description",
                "Best Life Domains": "bestLifeDomains",
                "Best Adventure Archetypes": "bestAdventureArchetypes",
            },
            {
                "Best Life Domains": split_multi,
                "Best Adventure Archetypes": split_multi,
            },
        ),
        "adventureArchetypes": parse_named_rows(
            wb["Adventure Archetype Library"],
            "Adventure Archetype",
            "Adventure Archetype",
            {
                "Simple Description": "description",
                "Best Adventure Triggers": "bestAdventureTriggers",
                "Best Story Structures": "bestStoryStructures",
                "Best Emotional Arcs": "bestEmotionalArcs",
                "Best Story Worlds": "bestStoryWorlds",
                "Notes": "notes",
            },
            {
                "Best Adventure Triggers": split_multi,
                "Best Story Structures": split_multi,
                "Best Emotional Arcs": split_multi,
                "Best Story Worlds": split_multi,
            },
        ),
        "adventureTriggers": parse_adventure_triggers(wb["Adventure Trigger Library"]),
        "storyWorlds": parse_named_rows(
            wb["Story World Library"],
            "Story World",
            "Story World",
            {"Best Adventure Archetypes": "bestAdventureArchetypes"},
            {"Best Adventure Archetypes": split_multi},
        ),
        "storyWorldPacks": parse_named_rows(
            wb["Story World Packs"],
            "Story World",
            "Story World",
            {
                "Theme": "theme",
                "Main Characters": "mainCharacters",
                "Supporting Characters": "supportingCharacters",
                "Preferred Character Types": "preferredCharacterTypes",
                "Settings": "settings",
                "Nature Mechanics": "natureMechanics",
                "Props": "props",
                "Typical Missions": "typicalMissions",
                "Typical Obstacles": "typicalObstacles",
                "Best Life Domains": "bestLifeDomains",
                "Best Adventure Archetypes": "bestAdventureArchetypes",
                "Best Emotional Arcs": "bestEmotionalArcs",
                "Best Ganesha Wisdom Elements": "bestWisdomElements",
                "Notes": "notes",
            },
            {
                "Main Characters": split_multi,
                "Supporting Characters": split_multi,
                "Preferred Character Types": split_multi,
                "Settings": split_multi,
                "Nature Mechanics": split_multi,
                "Props": split_multi,
                "Typical Missions": split_multi,
                "Typical Obstacles": split_multi,
                "Best Life Domains": split_multi,
                "Best Adventure Archetypes": split_multi,
                "Best Emotional Arcs": split_multi,
                "Best Ganesha Wisdom Elements": split_multi,
            },
        ),
        "missions": parse_named_rows(
            wb["Mission Library"],
            "Mission",
            "Mission",
            {
                "Mission Group": "missionGroup",
                "Simple Description": "description",
                "Best Life Domains": "bestLifeDomains",
                "Best Story Actions": "bestStoryActions",
                "Best Story Worlds": "bestStoryWorlds",
                "Best Adventure Archetypes": "bestAdventureArchetypes",
                "Best Core Needs": "bestCoreNeeds",
                "Difficulty": "difficulty",
                "Notes": "notes",
            },
            {
                "Best Life Domains": split_multi,
                "Best Story Actions": split_multi,
                "Best Story Worlds": split_multi,
                "Best Adventure Archetypes": split_multi,
                "Best Core Needs": split_multi,
            },
        ),
        "obstacles": parse_named_rows(
            wb["Obstacle Library"],
            "Obstacle",
            "Obstacle",
            {
                "Obstacle Group": "obstacleGroup",
                "Type": "type",
                "Simple Description": "description",
                "Best Life Domains": "bestLifeDomains",
                "Best Story Actions": "bestStoryActions",
                "Best Story Worlds": "bestStoryWorlds",
                "Escalation Type": "escalationType",
                "Difficulty": "difficulty",
                "Notes": "notes",
            },
            {
                "Best Life Domains": split_multi,
                "Best Story Actions": split_multi,
                "Best Story Worlds": split_multi,
            },
        ),
        "storyStructures": parse_named_rows(
            wb["Story Structure Library (100)"],
            "#",
            "Story Structure",
            {
                "#": "number",
                "Core Pattern": "corePattern",
                "Simple Description": "description",
                "Why It Works": "whyItWorks",
                "Best Age": "bestAge",
                "Scene Flow Template": "sceneFlowTemplate",
                "Best Adventure Archetypes": "bestAdventureArchetypes",
                "Best Opening": "bestOpening",
                "Best Ending": "bestEnding",
                "Best Replay Hooks": "bestReplayHooks",
                "Best Read-Aloud Devices": "bestReadAloudDevices",
            },
            {
                "Best Adventure Archetypes": split_multi,
                "Best Replay Hooks": split_multi,
                "Best Read-Aloud Devices": split_multi,
            },
        ),
        "emotionalArcs": parse_named_rows(
            wb["Emotional Arc Library"],
            "Emotional Arc",
            "Emotional Arc",
            {
                "Emotional Journey": "journey",
                "Simple Description": "description",
                "Why It Works": "whyItWorks",
                "When to Use": "whenToUse",
                "Best Age": "bestAge",
                "Emotional Building Blocks": "buildingBlocks",
                "Avoid": "avoid",
                "Short Form": "shortForm",
            },
            {"Emotional Building Blocks": split_multi, "Emotional Journey": split_arrows},
        ),
        "openings": parse_named_rows(
            wb["Opening Library"],
            "Opening Type",
            "Opening Type",
            {
                "Simple Description": "description",
                "How It Works": "howItWorks",
                "Why It Works": "whyItWorks",
                "Best Story Structures": "bestStoryStructures",
                "Best Adventure Archetypes": "bestAdventureArchetypes",
                "Best Emotional Arcs": "bestEmotionalArcs",
                "Best Story Types": "bestStoryTypes",
                "Best Age": "bestAge",
                "Opening Formula": "openingFormula",
                "Opening Emotion": "openingEmotion",
                "Illustration Goal": "illustrationGoal",
                "Page-Turn Hook": "pageTurnHook",
                "Writing Tips": "writingTips",
            },
            {
                "Best Story Structures": split_multi,
                "Best Adventure Archetypes": split_multi,
                "Best Emotional Arcs": split_multi,
                "Best Story Types": split_multi,
            },
        ),
        "endings": parse_named_rows(
            wb["Ending Library"],
            "Ending Type",
            "Ending Type",
            {
                "Simple Description": "description",
                "How It Works": "howItWorks",
                "Why It Works": "whyItWorks",
                "Best Story Structures": "bestStoryStructures",
                "Best Adventure Archetypes": "bestAdventureArchetypes",
                "Best Emotional Arcs": "bestEmotionalArcs",
                "Best Story Types": "bestStoryTypes",
                "Best Age": "bestAge",
                "Ending Formula": "endingFormula",
                "Emotional Resolution": "emotionalResolution",
                "Final Line Style": "finalLineStyle",
                "Replay Trigger": "replayTrigger",
                "Illustration Goal": "illustrationGoal",
                "Writing Tips": "writingTips",
            },
            {
                "Best Story Structures": split_multi,
                "Best Adventure Archetypes": split_multi,
                "Best Emotional Arcs": split_multi,
                "Best Story Types": split_multi,
            },
        ),
        "replayHooks": parse_named_rows(
            wb["Replay Hook Library"],
            "Replay Hook",
            "Replay Hook",
            {
                "Category": "category",
                "Simple Description": "description",
                "How It Works": "howItWorks",
                "Why It Works": "whyItWorks",
                "Complexity": "complexity",
                "Best Story Stage": "bestStoryStage",
                "Best Story Structures": "bestStoryStructures",
                "Best Adventure Archetypes": "bestAdventureArchetypes",
                "Best Emotional Arcs": "bestEmotionalArcs",
                "Best Age": "bestAge",
                "Notes": "notes",
            },
            {
                "Best Story Structures": split_multi,
                "Best Adventure Archetypes": split_multi,
                "Best Emotional Arcs": split_multi,
            },
        ),
        "readAloudDevices": parse_named_rows(
            wb["Read-Aloud Library"],
            "Read-Aloud Device",
            "Read-Aloud Device",
            {
                "Category": "category",
                "Simple Description": "description",
                "How It Works": "howItWorks",
                "Why It Works": "whyItWorks",
                "Best Story Stage": "bestStoryStage",
                "Best Story Structures": "bestStoryStructures",
                "Best Adventure Archetypes": "bestAdventureArchetypes",
                "Best Emotional Arcs": "bestEmotionalArcs",
                "Best Age": "bestAge",
                "Example": "example",
                "Notes": "notes",
            },
            {
                "Best Story Structures": split_multi,
                "Best Adventure Archetypes": split_multi,
                "Best Emotional Arcs": split_multi,
            },
        ),
        "titleFormulas": parse_named_rows(
            wb["Title Formula Library"],
            "Title Formula",
            "Title Formula",
            {
                "Formula Pattern": "pattern",
                "Simple Description": "description",
                "Why It Works": "whyItWorks",
                "Best Story Types": "bestStoryTypes",
                "Best Story Structures": "bestStoryStructures",
                "Best Adventure Archetypes": "bestAdventureArchetypes",
                "Best Emotional Arcs": "bestEmotionalArcs",
                "Best Age": "bestAge",
                "Example Titles": "exampleTitles",
                "Notes": "notes",
            },
            {
                "Best Story Types": split_multi,
                "Best Story Structures": split_multi,
                "Best Adventure Archetypes": split_multi,
                "Best Emotional Arcs": split_multi,
                "Example Titles": split_multi,
            },
        ),
        "storyConflicts": parse_named_rows(
            wb["Story Conflict Library"],
            "Story Conflict",
            "Story Conflict",
            {
                "Simple Description": "description",
                "Generated From": "generatedFrom",
                "Best Story Structures": "bestStoryStructures",
                "Best Emotional Arcs": "bestEmotionalArcs",
            },
            {"Best Story Structures": split_multi, "Best Emotional Arcs": split_multi},
        ),
        "escalations": parse_named_rows(
            wb["Escalation Library"],
            "Escalation Type",
            "Escalation Type",
            {
                "Simple Description": "description",
                "Common Use": "commonUse",
            },
        ),
        "storyQualityScorecard": parse_scorecard(wb["Story Quality Scorecard"]),
    }

    return {
        "sourceWorkbook": str(workbook_path),
        "libraryCount": len(libraries),
        "libraries": libraries,
    }


def main():
    if len(sys.argv) != 3:
        raise SystemExit("Usage: extract_prana_story_libraries.py <input.xlsx> <output_dir>")

    workbook_path = Path(sys.argv[1]).expanduser().resolve()
    output_dir = Path(sys.argv[2]).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    bundle = build_bundle(workbook_path)

    bundle_path = output_dir / "library-bundle.json"
    bundle_path.write_text(json.dumps(bundle, ensure_ascii=False, indent=2), encoding="utf-8")
    (output_dir / "library-bundle.js").write_text(
        "window.PRANA_STORY_LIBRARY_BUNDLE = " + json.dumps(bundle, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )

    manifest = {"sourceWorkbook": str(workbook_path), "libraries": []}
    for name, items in bundle["libraries"].items():
        file_name = f"{name}.json"
        (output_dir / file_name).write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
        manifest["libraries"].append({"name": name, "file": file_name, "count": len(items)})

    (output_dir / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
